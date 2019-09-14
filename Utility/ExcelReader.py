import openpyxl

class ExcelReader() :

    testDataFilePath = ' '
    testDataSheet = ' '      # workbook.active

    def __init__(self,testDataFilePath,testDataSheet):
        self.testDataFilePath = testDataFilePath
        self.testDataSheet = testDataSheet
        workbook = openpyxl.load_workbook(self.testDataFilePath)
        self.testDataSheet = workbook[self.testDataSheet]

    def getData(self,testCaseName, keyName):
        total_rows = self.testDataSheet.max_row
        total_cols = self.testDataSheet.max_column

        value = ''

        for row in range(2, total_rows+1) :
            testName = self.testDataSheet.cell(row = row, column = 1).value
            col = 1
            colName = self.testDataSheet.cell(row = 1, column = col).value
            #print(testName)
            if(testName == testCaseName):

                #print(colName)
                while(colName is not None):
                    colName = self.testDataSheet.cell(row = 1, column = col).value
                    #print(colName)
                    if(colName == keyName):
                        value = self.testDataSheet.cell(row = row, column = col).value
                        #print(value)
                        break
                    col = col + 1
                break
        return value


    def getUserName(self,testCaseName):
        return self.getData(testCaseName, 'Username')

    def getPassword(self,testCaseName):
        return self.getData(testCaseName, 'Password')


'''
excelReader = ExcelReader("E:/Python+Selenium/Py_Charm_Edu_Wrkspace/FRAMEWORKS/UT_PyTest_20190909/DataProviders/TestData.xlsx", 'TestData')
print(excelReader.getData('LoginTest001','Username'))
print(excelReader.getData('LoginTest001','Password'))
print(excelReader.getData('LoginTest002','Username'))
print(excelReader.getData('LoginTest002','Password'))
print(excelReader.getUserName('LoginTest001'))
print(excelReader.getUserName('LoginTest002'))


excelReader = ExcelReader("E:/Python+Selenium/Py_Charm_Edu_Wrkspace/FRAMEWORKS/UT_PyTest_20190909/DataProviders/TestData.xlsx", 'TestData2')
print(excelReader.getData('TD02-LoginTest001','Username'))
print(excelReader.getData('TD02-LoginTest001','Password'))
print(excelReader.getData('TD02-LoginTest002','Username'))
print(excelReader.getData('TD02-LoginTest002','Password'))
print(excelReader.getUserName('TD02-LoginTest001'))
print(excelReader.getUserName('TD02-LoginTest002'))
'''
